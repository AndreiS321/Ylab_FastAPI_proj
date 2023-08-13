import os

import httplib2
from googleapiclient.discovery import build
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from admin.dataclasses import ExcelDish, ExcelMenu, ExcelSubmenu


def parse_excel(file: str, sheet_name: str = "Лист1") -> list[dict]:
    """Возвращает список ExcelMenus в виде list[dict]"""
    excel_path = os.path.join(os.path.dirname(__file__), file)
    book: Workbook = load_workbook(excel_path)
    sheet: Worksheet = book.get_sheet_by_name(sheet_name)
    res = []
    cur = 1
    while True:
        menu_name = sheet.cell(cur, 2).value
        menu_descr = sheet.cell(cur, 3).value
        if isinstance(menu_name, str) and isinstance(menu_descr, str):
            cur += 1
            cur_menu = ExcelMenu(title=menu_name, description=menu_descr, submenus=[])
            res.append(cur_menu)
            while True:
                submenu_name = sheet.cell(cur, 3).value
                submenu_descr = sheet.cell(cur, 4).value
                if isinstance(submenu_name, str) and isinstance(submenu_descr, str):
                    cur += 1
                    cur_submenu = ExcelSubmenu(
                        title=submenu_name, description=submenu_descr, dishes=[]
                    )
                    cur_menu.submenus.append(cur_submenu)
                    while True:
                        dish_name = sheet.cell(cur, 4).value
                        dish_descr = sheet.cell(cur, 5).value
                        dish_price = sheet.cell(cur, 6).value
                        dish_discount = sheet.cell(cur, 7).value
                        if dish_name and dish_descr and dish_price:
                            dish_price = (
                                dish_price * (1 - dish_discount)
                                if dish_discount and (0 < dish_discount < 1)
                                else dish_price
                            )
                            cur += 1
                            cur_submenu.dishes.append(
                                ExcelDish(
                                    title=dish_name,
                                    description=dish_descr,
                                    price=dish_price,
                                )
                            )
                        else:
                            break
                else:
                    break
        else:
            break
    res_dict = [menu.as_dict for menu in res]
    return res_dict


def parse_google_sheet(
        key_file: str, sheet_id: str, sheet: str = "Лист1"
) -> list[dict]:
    """Возвращает список ExcelMenus в виде list[dict]"""
    creds_json = os.path.join(os.path.dirname(__file__), key_file)
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]

    creds_service = ServiceAccountCredentials.from_json_keyfile_name(
        creds_json, scopes
    ).authorize(httplib2.Http())
    cli = build("sheets", "v4", http=creds_service)

    sheet = (
        cli.spreadsheets()
        .values()
        .get(spreadsheetId=sheet_id, range=f"{sheet}!A1:G999",valueRenderOption="UNFORMATTED_VALUE")
        .execute()
    )
    sheet = sheet.get("values",[])
    res = []
    cur = 0
    while True:
        try:
            menu_name = sheet[cur][1]
            menu_descr = sheet[cur][2]
        except IndexError:
            break
        if (
                menu_name
                and menu_descr
                and (isinstance(menu_name,str))
                and (isinstance(menu_descr,str))
        ):
            cur += 1
            cur_menu = ExcelMenu(title=menu_name, description=menu_descr, submenus=[])
            res.append(cur_menu)
            while True:
                try:
                    submenu_name: str = sheet[cur][2]
                    submenu_descr = sheet[cur][3]
                except IndexError:
                    break
                if (
                        submenu_name
                        and submenu_descr
                        and (isinstance(submenu_name,str))
                        and (isinstance(submenu_descr,str))
                ):
                    cur += 1
                    cur_submenu = ExcelSubmenu(
                        title=submenu_name, description=submenu_descr, dishes=[]
                    )
                    cur_menu.submenus.append(cur_submenu)
                    while True:
                        try:
                            dish_name = sheet[cur][3]
                            dish_descr = sheet[cur][4]
                            dish_price = sheet[cur][5]
                        except IndexError:
                            break
                        try:
                            dish_discount = sheet[cur][6]
                        except IndexError:
                            dish_discount = 0
                        if dish_name and dish_descr and dish_price:
                            dish_price = (
                                dish_price * (1 - dish_discount)
                                if dish_discount and (0 < dish_discount < 1)
                                else dish_price
                            )
                            cur += 1
                            cur_submenu.dishes.append(
                                ExcelDish(
                                    title=dish_name,
                                    description=dish_descr,
                                    price=dish_price,
                                )
                            )
                        else:
                            break
                else:
                    break
        else:
            break
    res_dict = [menu.as_dict for menu in res]
    return res_dict
